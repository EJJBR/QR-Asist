"""
QR-Asist - Servidor Flask
Rutas y endpoints de la aplicación
"""

from flask import Flask, render_template, request, jsonify, send_file
import os
import json
from datetime import datetime
import csv

# Importar módulos del proyecto
from modules.generador_qr import GeneradorQR
from modules.lector_qr import LectorQR
from modules.consolidador import Consolidador
import config

app = Flask(__name__)
app.config['SECRET_KEY'] = 'qr-asist-secret-key-2026'

# Configuración de rutas
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATOS_DIR = os.path.join(BASE_DIR, 'datos')
QR_DIR = os.path.join(DATOS_DIR, 'qr_codes')
REGISTROS_DIR = os.path.join(BASE_DIR, 'registro')
REPORTES_DIR = os.path.join(BASE_DIR, 'reportes')

# Asegurar que existan las carpetas
os.makedirs(QR_DIR, exist_ok=True)
os.makedirs(REGISTROS_DIR, exist_ok=True)
os.makedirs(REPORTES_DIR, exist_ok=True)

# Instanciar módulos
generador = GeneradorQR()
lector = LectorQR(laptop_id=config.LAPTOP_ID)
consolidador = Consolidador(
    carpeta_compartida=config.CARPETA_COMPARTIDA,
    carpeta_consolidados=config.CONSOLIDADOS_DIR
)

# ==================== RUTAS DE PÁGINAS ====================

@app.route('/')
def index():
    """Página principal"""
    return render_template('index.html')

@app.route('/generar-qr')
def generar_qr_page():
    """Módulo de generación de códigos QR"""
    return render_template('generar_qr.html')

@app.route('/leer-qr')
def leer_qr_page():
    """Módulo de lectura de QR (próximamente)"""
    return render_template('leer_qr.html')

@app.route('/consolidar')
def consolidar_page():
    """Módulo de consolidación (próximamente)"""
    return render_template('consolidar.html')

# ==================== API ENDPOINTS ====================

@app.route('/api/cargar-archivo', methods=['POST'])
def cargar_archivo():
    """Cargar y leer archivo CSV o Excel de alumnos"""
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'No se envió ningún archivo'}), 400
        
        archivo = request.files['archivo']
        
        if archivo.filename == '':
            return jsonify({'error': 'Nombre de archivo vacío'}), 400
        
        # Validar extensión
        if not (archivo.filename.endswith('.csv') or archivo.filename.endswith('.xlsx')):
            return jsonify({'error': 'El archivo debe ser CSV o Excel (.xlsx)'}), 400
        
        alumnos = []
        
        # Procesar según el tipo de archivo
        if archivo.filename.endswith('.csv'):
            # Leer contenido del CSV
            contenido = archivo.read().decode('utf-8')
            lineas = contenido.strip().split('\n')
            
            # Parsear CSV
            reader = csv.DictReader(lineas)
            
            for fila in reader:
                if 'ID' in fila and 'NOMBRE_COMPLETO' in fila:
                    alumnos.append({
                        'id': fila['ID'].strip(),
                        'nombre': fila['NOMBRE_COMPLETO'].strip()
                    })
        
        elif archivo.filename.endswith('.xlsx'):
            # Importar openpyxl
            from openpyxl import load_workbook
            
            # Cargar el archivo Excel
            wb = load_workbook(archivo)
            ws = wb.active
            
            # Leer encabezados (primera fila)
            encabezados = []
            for cell in ws[1]:
                encabezados.append(cell.value)
            
            # Buscar índices de las columnas ID y NOMBRE_COMPLETO
            try:
                idx_id = encabezados.index('ID')
                idx_nombre = encabezados.index('NOMBRE_COMPLETO')
            except ValueError:
                return jsonify({'error': 'El archivo Excel debe tener columnas ID y NOMBRE_COMPLETO'}), 400
            
            # Leer datos (desde la fila 2 en adelante)
            for fila in ws.iter_rows(min_row=2, values_only=True):
                if fila[idx_id] and fila[idx_nombre]:
                    alumnos.append({
                        'id': str(fila[idx_id]).strip(),
                        'nombre': str(fila[idx_nombre]).strip()
                    })
        
        return jsonify({
            'success': True,
            'alumnos': alumnos,
            'total': len(alumnos)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/generar-qr', methods=['POST'])
def generar_qr_api():
    """Generar códigos QR"""
    try:
        datos = request.json
        
        nivel = datos.get('nivel')
        grado = datos.get('grado')
        seccion = datos.get('seccion')
        alumnos = datos.get('alumnos', [])
        
        if not all([nivel, grado, seccion, alumnos]):
            return jsonify({'error': 'Faltan datos requeridos'}), 400
        
        # Generar QR para cada alumno
        resultado = generador.generar_codigos_qr(
            alumnos=alumnos,
            nivel=nivel,
            grado=grado,
            seccion=seccion
        )
        
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/generar-pdf', methods=['POST'])
def generar_pdf_api():
    """Generar PDF con códigos QR"""
    try:
        datos = request.json
        
        nivel = datos.get('nivel')
        grado = datos.get('grado')
        seccion = datos.get('seccion')
        alumnos = datos.get('alumnos', [])
        
        # Generar PDF
        pdf_path = generador.crear_pdf_impresion(
            alumnos=alumnos,
            nivel=nivel,
            grado=grado,
            seccion=seccion
        )
        
        if pdf_path and os.path.exists(pdf_path):
            return send_file(
                pdf_path,
                as_attachment=True,
                download_name=os.path.basename(pdf_path)
            )
        else:
            return jsonify({'error': 'No se pudo generar el PDF'}), 500
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/agregar-alumno', methods=['POST'])
def agregar_alumno_api():
    """Agregar un alumno individual"""
    try:
        datos = request.json
        
        id_alumno = datos.get('id')
        nombre = datos.get('nombre')
        nivel = datos.get('nivel')
        grado = datos.get('grado')
        seccion = datos.get('seccion')
        
        if not all([id_alumno, nombre, nivel, grado, seccion]):
            return jsonify({'error': 'Faltan datos del alumno'}), 400
        
        # Generar QR para este alumno
        alumno = {'id': id_alumno, 'nombre': nombre}
        
        resultado = generador.generar_qr_individual(
            alumno=alumno,
            nivel=nivel,
            grado=grado,
            seccion=seccion
        )
        
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== API MÓDULO 2: LECTOR QR ====================

@app.route('/api/registrar-asistencia', methods=['POST'])
def registrar_asistencia_api():
    """Registrar asistencia desde QR escaneado"""
    try:
        datos = request.json
        datos_qr = datos.get('qr_data')
        
        if not datos_qr:
            return jsonify({'error': 'No se recibieron datos del QR'}), 400
        
        resultado = lector.registrar_asistencia(datos_qr)
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/estadisticas-hoy', methods=['GET'])
def estadisticas_hoy_api():
    """Obtener estadísticas del día actual"""
    try:
        total = lector.contar_registros_hoy()
        ultimos = lector.obtener_ultimos_registros(5)
        
        return jsonify({
            'success': True,
            'total_hoy': total,
            'ultimos_registros': ultimos
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/listar-archivos', methods=['GET'])
def listar_archivos_api():
    """Listar archivos de registro disponibles"""
    try:
        archivos = lector.listar_archivos_registro()
        
        return jsonify({
            'success': True,
            'archivos': archivos,
            'total': len(archivos)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/enviar-archivos', methods=['POST'])
def enviar_archivos_api():
    """Enviar archivos seleccionados a PC central"""
    try:
        datos = request.json
        archivos = datos.get('archivos', [])
        
        if not archivos:
            return jsonify({'error': 'No se seleccionaron archivos'}), 400
        
        # Obtener carpeta de destino desde config
        carpeta_destino = config.CARPETA_COMPARTIDA
        
        # Enviar archivos
        resultado = lector.enviar_multiples(archivos, carpeta_destino)
        
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/verificar-red', methods=['GET'])
def verificar_red_api():
    """Verificar si la carpeta de red está disponible"""
    try:
        carpeta_destino = config.CARPETA_COMPARTIDA
        disponible = os.path.exists(carpeta_destino)
        
        return jsonify({
            'success': True,
            'disponible': disponible,
            'ruta': carpeta_destino
        })
    
    except Exception as e:
        return jsonify({
            'success': True,
            'disponible': False,
            'error': str(e)
        })

# ==================== API MÓDULO 3: CONSOLIDADOR ====================

@app.route('/api/consolidar-todo', methods=['POST'])
def consolidar_todo_api():
    """Ejecutar consolidación de todos los archivos"""
    try:
        datos = request.json or {}
        forzar = datos.get('forzar', False)
        
        resultado = consolidador.consolidar_todo(forzar=forzar)
        return jsonify(resultado)
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/cargar-consolidados', methods=['POST'])
def cargar_consolidados_api():
    """Cargar consolidados en un rango de fechas"""
    try:
        datos = request.json
        fecha_inicio_str = datos.get('fechaInicio')
        fecha_fin_str = datos.get('fechaFin')
        
        if not fecha_inicio_str or not fecha_fin_str:
            return jsonify({'error': 'Faltan fechas'}), 400
        
        # Parsear fechas
        fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d")
        fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d")
        
        # Cargar registros
        registros = consolidador.cargar_consolidados(fecha_inicio, fecha_fin)
        
        return jsonify({
            'success': True,
            'registros': registros,
            'total': len(registros)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/listar-consolidados', methods=['GET'])
def listar_consolidados_api():
    """Listar todos los consolidados disponibles"""
    try:
        consolidados = consolidador.listar_consolidados()
        
        return jsonify({
            'success': True,
            'consolidados': consolidados,
            'total': len(consolidados)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/exportar-reporte', methods=['POST'])
def exportar_reporte_api():
    """Exportar reporte filtrado a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        datos = request.json
        registros = datos.get('registros', [])
        
        if not registros:
            return jsonify({'error': 'No hay datos para exportar'}), 400
        
        # Crear libro Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Asistencias"
        
        # Encabezados
        headers = ['ID', 'Nombre Completo', 'Nivel', 'Grado', 'Sección', 
                  'Fecha', 'Hora', 'Laptop']
        ws.append(headers)
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Datos
        for registro in registros:
            ws.append([
                registro.get('ID', ''),
                registro.get('NOMBRE_COMPLETO', ''),
                registro.get('NIVEL', ''),
                registro.get('GRADO', ''),
                registro.get('SECCION', ''),
                registro.get('FECHA', ''),
                registro.get('HORA', ''),
                registro.get('LAPTOP', '')
            ])
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        
        # Guardar en memoria (BytesIO) en lugar de disco
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"reporte_filtrado_{timestamp}.xlsx"
        
        # Enviar archivo directamente desde memoria
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== MANEJO DE ERRORES ====================

@app.errorhandler(404)
def page_not_found(e):
    """Página no encontrada"""
    return render_template('index.html'), 404

@app.errorhandler(500)
def internal_error(e):
    """Error interno del servidor"""
    return jsonify({'error': 'Error interno del servidor'}), 500

if __name__ == '__main__':
    app.run(debug=True)